import pandas as pd

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from pypinyin import lazy_pinyin, Style
except ImportError:
    lazy_pinyin = None

# 步骤2: 读取Excel文件的所有工作表
excel_file = pd.ExcelFile('example1.xlsx')
sheet_names = excel_file.sheet_names

# 目标列名称
possible_columns = ['Name', '姓名']

# 使用openpyxl在前五行中查找目标列名称
if openpyxl is None:
    raise ImportError('未安装 openpyxl，无法查找目标列。请运行: pip install openpyxl')
wb = openpyxl.load_workbook('example1.xlsx')
dfs = {}
for sheet_name in sheet_names:
    ws = wb[sheet_name]
    target_col_index = None
    for row in range(1, 6):  # 前五行
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value in possible_columns:
                target_col_index = col
                break
        if target_col_index:
            break
    
    df = pd.read_excel('example1.xlsx', sheet_name=sheet_name, header=0)
    dfs[sheet_name] = (df, target_col_index)

def convert_to_pinyin_initials(value):
    if not isinstance(value, str) or len(value) == 0:
        return value
    if lazy_pinyin is None:
        raise ImportError(
            '未安装 pypinyin，无法将汉字转换为拼音首字母。请运行: pip install pypinyin'
        )
    initials = lazy_pinyin(value, style=Style.FIRST_LETTER, errors='default')
    return ''.join(item[0].upper() for item in initials if item)

# 处理每个工作表
for sheet_name, (df, target_col_index) in dfs.items():
    if target_col_index is None:
        print(f"警告: 工作表 '{sheet_name}' 前五行中未找到目标列 {possible_columns}，跳过。")
        continue

    target_column = df.columns[target_col_index - 1]  # openpyxl 列从1开始，df.columns从0
    if target_col_index - 1 >= len(df.columns):
        print(f"警告: 工作表 '{sheet_name}' 列索引 {target_col_index} 超出范围，跳过。")
        continue

    new_column = f"{target_column}_pinyin"
    if new_column in df.columns:
        print(f"警告: 工作表 '{sheet_name}' 中 '{new_column}' 列已存在，跳过。")
        continue

    insert_position = df.columns.get_loc(target_column)
    pinyin_values = df[target_column].apply(convert_to_pinyin_initials)
    df.insert(insert_position, new_column, pinyin_values)
    print(f"工作表 '{sheet_name}' 的 '{target_column}' 列已转换为拼音首字母。")

# 步骤4: 写回Excel文件
with pd.ExcelWriter('modified_example.xlsx') as writer:
    for sheet_name, (df, _) in dfs.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)

