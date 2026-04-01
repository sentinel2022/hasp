import os
from flask import Flask, render_template, request, redirect, url_for, flash, session
import pandas as pd
from werkzeug.utils import secure_filename
import sys
import openpyxl

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# 配置上传文件夹
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_uploaded_files():
    """获取已上传的文件列表"""
    files = []
    for filename in os.listdir(app.config['UPLOAD_FOLDER']):
        if allowed_file(filename):
            try:
                # 尝试解码文件名（针对Python 2兼容性）
                if sys.version_info[0] < 3:
                    filename = filename.decode('utf-8')
            except:
                pass
            files.append(filename)
    return files

def secure_filename_custom(filename):
    """自定义安全文件名函数，保留中文字符"""
    # 保留中文字符、数字、字母、下划线和点
    keep_chars = (' ', '.', '_', '-')
    filename = "".join(c for c in filename if c.isalnum() or c in keep_chars).strip()
    return filename

def search_in_excel(filepath, keyword):
    """在Excel文件中搜索关键字"""
    results = []
    
    try:
        # 使用openpyxl直接读取，提高速度
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 获取标题行（第一行）
            headers = []
            for cell in ws[1]:
                headers.append(cell.value if cell.value is not None else '')
            
            matched_rows = []
            
            # 逐行读取，从第二行开始（跳过标题）
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # 检查行中是否有单元格包含关键字
                row_contains = any(str(cell).lower().find(keyword.lower()) != -1 for cell in row if cell is not None)
                
                if row_contains:
                    # 将行转换为字典
                    row_dict = {}
                    for col_idx, cell in enumerate(row):
                        header = headers[col_idx] if col_idx < len(headers) else f'Column_{col_idx+1}'
                        row_dict[header] = cell
                    matched_rows.append(row_dict)
            
            if matched_rows:
                results.append({
                    'sheet_name': sheet_name,
                    'headers': headers,
                    'data': matched_rows
                })
        
        wb.close()
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        flash('读取Excel文件时出错，请检查文件格式')
    
    return results

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # 获取表单数据
        keyword = request.form.get('keyword', '').strip()
        file_option = request.form.get('file_option', 'existing')
        selected_file = request.form.get('selected_file', '')
        
        # 检查关键字
        if not keyword:
            flash('请输入搜索关键字')
            return redirect(request.url)
        
        # 处理文件选择
        file_to_search = None
        
        if file_option == 'new':
            # 处理新文件上传
            if 'file' not in request.files:
                flash('未选择文件')
                return redirect(request.url)
                
            file = request.files['file']
            if file.filename == '':
                flash('未选择文件')
                return redirect(request.url)
                
            if file and allowed_file(file.filename):
                # 处理中文文件名
                filename = secure_filename_custom(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                file_to_search = filename
                session['selected_file'] = filename
                flash('文件上传成功')
            else:
                flash('只允许上传Excel文件 (.xlsx, .xls)')
                return redirect(request.url)
        else:
            # 处理已存在的文件选择
            if not selected_file:
                flash('请选择一个文件')
                return redirect(request.url)
                
            file_to_search = selected_file
            session['selected_file'] = selected_file
        
        # 搜索关键字
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file_to_search)
        results = search_in_excel(filepath, keyword)
        
        if not results:
            flash(f'未找到包含 "{keyword}" 的内容')
            return redirect(request.url)
        
        return render_template('results.html', 
                             keyword=keyword, 
                             filename=file_to_search, 
                             results=results)
    
    # GET请求 - 显示表单
    uploaded_files = get_uploaded_files()
    return render_template('index.html', 
                         uploaded_files=uploaded_files,
                         selected_file=session.get('selected_file', ''))

# 确保模板渲染时使用正确的编码
@app.after_request
def add_header(response):
    response.headers['Content-Type'] = 'text/html; charset=utf-8'
    return response

if __name__ == '__main__':
    app.run(debug=True, port=5001)